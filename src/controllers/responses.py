from openpyxl import load_workbook, Workbook
import os
from kivy.logger import Logger

from utils.conversion import Conversion
from config import Config


class ResponsesControllerState:

    def __init__(self, config, workbook=None, request_response_pairs={},
                 sniffer_stack=[]):
        self.workbook = workbook
        self.request_response_pairs = request_response_pairs
        self.sniffer_stack = sniffer_stack
        self.config = config


class ResponsesController:

    def __init__(self, state):
        self.state = state

    def run(self):
        self.load_pairs()
        return True

    def load_pairs(self):
        config = self.state.config
        filename = config["FILENAME"]

        if (os.path.isfile(filename)):
            self.state.workbook = load_workbook(filename)
            sheet = self.state.workbook.active
            index = config["START_ROW"]
            while True:
                req = sheet[config["REQUESTS"]["COLUMN"]
                            + str(index)].value
                res = sheet[config["RESPONSES"]["COLUMN"]
                            + str(index)].value
                if (not (req and res)):
                    break
                self.state.request_response_pairs[str(req)] = str(res)
                index += 1
        return True

    def set_pair(self, req, res):
        req_string = Conversion.bytes_to_hex_string(req)
        res_string = Conversion.bytes_to_hex_string(res)

        self.state.request_response_pairs[req_string] = res_string
        return True

    def get_response(self, req):
        req_string = Conversion.bytes_to_hex_string(req)

        res = Conversion.hex_string_to_bytes(
                self.state.request_response_pairs[req_string]) \
            if req_string in self.state.request_response_pairs \
            else Config.ACK
        return res

    def save_pairs(self):
        config = self.state.config
        filename = config["FILENAME"]

        self.state.workbook = Workbook()
        sheet = self.state.workbook.active
        req_column = config["REQUESTS"]["COLUMN"]
        res_column = config["RESPONSES"]["COLUMN"]
        sheet[req_column + "1"] = config["REQUESTS"]["TITLE"]
        sheet[res_column + "1"] = config["RESPONSES"]["TITLE"]

        index = config["START_ROW"]
        for req, res in self.state.request_response_pairs.items():
            sheet[req_column + str(index)] = req
            sheet[res_column + str(index)] = res
            index += 1

        self.state.workbook.save(filename)
